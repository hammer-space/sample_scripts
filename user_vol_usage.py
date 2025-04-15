#!/usr/bin/env python3
# Script to query Hammerspace metadata via HSTK then parse the JSON result into an Excel workbook.

# 07/28/24 peter@hammerspace.com

# Disclaimer: This script is provided "as is" without warranty of any kind, express or implied.
# Use at your own risk. The author is not liable for any damages or issues arising from its use.

# This script requires openpyxl, argparse, and hstk.  You can install with
#  pip3 install openpyxl
#  pip3 install argparse
#  pip3 install hstk

# TODO
# Check that HSTK is installed and working [done?]
# Path handling from command line parameter or cwd
#   Validate that path given is a mount from Hammerspace [done]
#   Convert "." to a path. [done]
#   Convert mount point to exported share.
# Show the mount / net use info for the path [partial]
# Output file name parameter
#   Check/handle output file already exists?
#   Append .xlsx if not included in output file name?
# Error handling of HSTK error (user insufficient permissions?)
# Put date/time on the report [done]

# Figure out how to parse hstk output into rows and cells [done - whew!]
# Figure out how to write .xlsx files [done]
#   Title (top of sheet and tab)  [done]
#   Column headings  [done]
#     Autofilter  [done]
#     Bottom border [done]
#   Column widths appropriate for data in fields [done]

import subprocess
import json
import openpyxl
from openpyxl.styles import Border, Side, Font
import argparse
import os.path
import datetime

# Parse command line arguments
parse = argparse.ArgumentParser()
parse.add_argument("scanpath", type=str, default=".")
parse.add_argument("xlfile", type=str)
#parse.add_argument("xlfile", type=str, required=True)
args=parse.parse_args()
#print ("Scanpath: ", args.scanpath)
#print ("xlfile: ", args.xlfile)
scanpath = args.scanpath
xlfile = args.xlfile

#print('os.path.abspath:  ', os.path.abspath(scanpath))
# os.path.realpath is useful for converting "." to a path.  In Windows it yields UNC path.
#print('os.path.realpath:  ', os.path.realpath(scanpath))

cmd='hs'
result = subprocess.run(cmd, shell = True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
#print('returncode: ' + str(result.returncode))
if result.returncode > 0:
  print("Hammerspace ToolKit (HSTK) is required and not found.")
  print("See https://github.com/hammer-space/hstk")
  print("or just do ")
  print("pip3 install hstk")
  print("You might also need a modern Python and pip3.")
  exit(2)

# From https://stackoverflow.com/a/40935194
def as_text(value):
    if value is None:
        return ""
    return str(value)

# Validate that scanpath is on Hammerspace.  Check scanpath?.attribute=inode_info is readable.
# This doesn't work in Windows because ? is special.
#if not os.path.exists(os.path.join(scanpath, '?.attribute=inode_info')):
#  print(scanpath + " invalid. Scanpath parameter must point to a directory on a Hammerspace share.")
#  exit(1)
# Instead, we'll use hs to get the ITEM_TYPE.  If scanpath is not on Hammerspace, we won't get an item type back.
cmd='hs eval -e type ' + scanpath
result = subprocess.run(cmd, shell = True, stdout=subprocess.PIPE)
typecheck = str(result.stdout)
try:
  typecheck.index("ITEM_TYPE")
except:
  print(scanpath + " invalid. Scanpath parameter must point to a directory on a Hammerspace share.")
  # But technically, a file would work too.
  exit(1)

# Run the hstk command to get the usage report metadata
cmd='hs --json sum -e        "IS_FILE?SUMS_TABLE{|::KEY={OWNER, OWNER_GROUP,INSTANCES[PARENT.ROW].VOLUME},|::VALUE={1FILE/files,SPACE_USED/bytes}}[ROWS(INSTANCES)]" ' + scanpath
#This one only counts files over a certain size.  (Customer request.)
#cmd='hs --json sum -e "SIZE > 1*MBYTE?SUMS_TABLE{|::KEY={OWNER, OWNER_GROUP,INSTANCES[PARENT.ROW].VOLUME},|::VALUE={1FILE/files,SPACE_USED/bytes}}[ROWS(INSTANCES)]" ' + scanpath
result = subprocess.run(cmd, shell = True, stdout=subprocess.PIPE)

wb = openpyxl.Workbook()
sheet = wb.active
# Sheet title
sheet.title = 'user-vol'
# os.path.realpath is useful for converting "." to a path.  In Windows it yields UNC path.
sheet.append(["", "Storage usage report for " + os.path.realpath(scanpath)])
sheet.append(["", datetime.datetime.now().strftime("%m/%d/%Y %H:%M:%S")])
sheet.append(["", ""])
sheet['B1'].font = Font(size=14,bold=True)
sheet.append(["User","Group","Storage Volume","Files","Bytes"])
for col in ('A', 'B', 'C', 'D', 'E'):
  sheet[col + '4'].border = Border(bottom=Side(border_style="medium"))


report = json.loads(result.stdout)
for line in report['SUMS_TABLE']:
  row = [
    line['KEY'][0][0]['HAMMERSCRIPT'].split("'")[1].split("|")[0],
    line['KEY'][0][1]['HAMMERSCRIPT'].split("'")[1],
    line['KEY'][0][2]['HAMMERSCRIPT'].split("'")[1],
    line['VALUE'][0][0],
    line['VALUE'][0][1]
    ]
#  print(row)
  sheet.append(row)

# Headings with autofilter
filters = sheet.auto_filter
filters.ref = "A4:E51"

#column widths 
# From https://stackoverflow.com/a/40935194
for column_cells in sheet.columns:
    length = max(len(as_text(cell.value)) for cell in column_cells)
#    print(openpyxl.utils.get_column_letter(column_cells[0].column))
#    print(length)
    sheet.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = length

wb.save(xlfile)
